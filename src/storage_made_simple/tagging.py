from __future__ import annotations

import csv
import io
import re
from typing import Iterable

SOURCE_PATTERNS: tuple[tuple[str, str], ...] = (
    (r"\birs\b", "from:irs"),
    (r"\bcitizens[\s-]?bank\b", "from:citizens-bank"),
    (r"\bfreedom[\s-]?mortgage\b", "from:freedom-mortgage"),
    (r"\bamerican[\s-]?express\b|\bamex\b", "from:american-express"),
)

TYPE_PATTERNS: tuple[tuple[str, str], ...] = (
    (r"\bstatement\b", "type:statement"),
    (r"\btransaction(?:s|\s+list)?\b", "type:transaction-list"),
    (r"\bsummary\b", "type:summary"),
    (r"\btax[\s-]?document\b", "type:tax-document"),
)

ACCOUNT_PATTERNS: tuple[tuple[str, str], ...] = (
    (r"\bchecking\b", "account:checking"),
    (r"\bsavings\b", "account:savings"),
    (r"\broth[\s-]?ira\b", "account:roth-ira"),
    (r"\bcredit[\s-]?card\b|\bcardmember\b", "account:credit-card"),
    (r"\bmortgage\b", "account:mortgage"),
)

FORM_PATTERNS: tuple[tuple[str, str], ...] = (
    (r"\bform[\s-]?1040\b|\b1040\b", "form:1040"),
    (r"\bform[\s-]?w[- ]?2\b|\bw[- ]?2\b", "form:w-2"),
    (r"\bform[\s-]?1099(?:[-\w]+)?\b", "form:1099"),
)

SCHEDULE_PATTERNS: tuple[tuple[str, str], ...] = (
    (r"\bschedule[\s-]?c\b", "schedule:c"),
    (r"\bschedule[\s-]?e\b", "schedule:e"),
)


def normalize_tag(value: str) -> str:
    raw = value.strip().lower()
    if not raw:
        return ""
    if ":" in raw:
        prefix, _, suffix = raw.partition(":")
        prefix = re.sub(r"[^a-z0-9]+", "-", prefix).strip("-")
        suffix = re.sub(r"[^a-z0-9]+", "-", suffix).strip("-")
        return f"{prefix}:{suffix}" if suffix else prefix
    return re.sub(r"[^a-z0-9]+", "-", raw).strip("-")


def dedupe_tags(tags: Iterable[str]) -> list[str]:
    deduped: list[str] = []
    seen: set[str] = set()
    for tag in tags:
        normalized = normalize_tag(tag)
        if not normalized or normalized in seen:
            continue
        deduped.append(normalized)
        seen.add(normalized)
    return deduped


def merge_tags(*tag_groups: Iterable[str]) -> list[str]:
    merged: list[str] = []
    for group in tag_groups:
        merged.extend(group)
    return dedupe_tags(merged)


def _derive_tags_from_text(text: str) -> list[str]:
    lowered = text.lower()
    tags: list[str] = []

    for pattern, tag in SOURCE_PATTERNS:
        if re.search(pattern, lowered, flags=re.IGNORECASE):
            tags.append(tag)

    for pattern, tag in TYPE_PATTERNS:
        if re.search(pattern, lowered, flags=re.IGNORECASE):
            tags.append(tag)

    for pattern, tag in ACCOUNT_PATTERNS:
        if re.search(pattern, lowered, flags=re.IGNORECASE):
            tags.append(tag)

    for pattern, tag in FORM_PATTERNS:
        if re.search(pattern, lowered, flags=re.IGNORECASE):
            tags.append(tag)

    for pattern, tag in SCHEDULE_PATTERNS:
        if re.search(pattern, lowered, flags=re.IGNORECASE):
            tags.append(tag)

    for year, month, day in re.findall(r"\b(20\d{2})-(0[1-9]|1[0-2])-(0[1-9]|[12]\d|3[01])\b", lowered):
        tags.append(f"date:{year}-{month}-{day}")
    for year, month in re.findall(r"\b(20\d{2})-(0[1-9]|1[0-2])\b", lowered):
        tags.append(f"date:{year}-{month}")
    for year in re.findall(r"\b(20\d{2})\b", lowered):
        tags.append(f"date:{year}")

    for year in re.findall(r"\btax[\s-]?year(?:\s|:|-)*(20\d{2})\b", lowered):
        tags.append(f"taxyear:{year}")

    for match in re.findall(r"\b(?:sub|subject)(?:\s|:|-)+([a-z0-9][a-z0-9\s-]{1,50})", lowered):
        tags.append(f"sub:{match.strip()}")

    return dedupe_tags(tags)


def _csv_signals(data: bytes) -> str:
    text = data.decode("utf-8-sig", errors="ignore")
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        return ""
    sample = rows[:5]
    flattened = [" ".join(cell for cell in row if cell) for row in sample]
    return "\n".join(flattened)


def _pdf_signals(data: bytes) -> str:
    try:
        from pypdf import PdfReader
    except Exception:
        return ""

    reader = PdfReader(io.BytesIO(data))
    texts: list[str] = []
    for index, page in enumerate(reader.pages):
        if index >= 3:
            break
        try:
            extracted = page.extract_text() or ""
        except Exception:
            extracted = ""
        if extracted:
            texts.append(extracted)
    return "\n".join(texts)


def _excel_signals(data: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except Exception:
        return ""

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    texts: list[str] = []
    for sheet in workbook.worksheets[:2]:
        for row in sheet.iter_rows(max_row=5, values_only=True):
            cells = [str(cell) for cell in row if cell is not None]
            if cells:
                texts.append(" ".join(cells))
    return "\n".join(texts)


def derive_tags_from_document(filename: str, data: bytes | None = None) -> list[str]:
    signals = [filename, filename.replace("_", " "), filename.replace("-", " ")]
    if data is not None:
        suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
        if suffix == "csv":
            signals.append(_csv_signals(data))
        elif suffix == "pdf":
            signals.append(_pdf_signals(data))
        elif suffix in {"xls", "xlsx"}:
            signals.append(_excel_signals(data))

    tags: list[str] = []
    for signal in signals:
        if signal:
            tags.extend(_derive_tags_from_text(signal))
    return dedupe_tags(tags)
